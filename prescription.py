#don't modify tis code

import cv2
import os
from datetime import date
from PIL import Image
import openpyxl as xl
today = date.today().strftime("%d/%m/%Y")

black = (1,1,1)
white = (255,255,255)

path = "values.xlsx"
db = xl.load_workbook(path)
table = db.active

sign2 = table.cell(row = 4, column = 2)
global serial
serial = table.cell(row = 6, column= 2)
global serial_val
serial_val = serial.value

global namelocx,agelocx,symlocx,diaglocx,genderlocx,medlocx,doselocx,timelocx,namelocy,agelocy,symlocy,diaglocy,genderlocy,medlocy

namelocx = table.cell(row =7 ,column=2)
namelocy= table.cell(row = 7,column=3)
agelocx= table.cell(row = 8,column=2)
agelocy= table.cell(row =8 ,column=3)
genderlocx= table.cell(row =9 ,column=2)
genderlocy= table.cell(row =9 ,column=3)
symlocx= table.cell(row = 10,column=2)
symlocy= table.cell(row = 10,column=3)
diaglocx= table.cell(row = 11,column=2)
diaglocy= table.cell(row = 11,column=3)
medlocx= table.cell(row = 12,column=2)
medlocy= table.cell(row = 12,column=3)
doselocx= table.cell(row = 13,column=2)
timelocx= table.cell(row = 14,column=2)

global addname,addage,addgender,addsignature,addsymptom,adddiagnosis,templete,addserial
addname = ""
addage = ""
addgender= ""
addsignature = ""
addsymptom = ""
adddiagnosis = ""
templete=""
addserial= serial_val

def serial_set():
    global serial, serial_val
    serial_val += 1
    serial.value = serial_val
    db.save("values.xlsx")

def pos(x):
    global medlocy
    num = medlocy.value + ((x-1)*40)
    return num

def generate():
    global addname,addage,addgender,addsignature,addsymptom,adddiagnosis,templete,namelocx,namelocy,agelocy,agelocx,symlocx,symlocy,diaglocy,diaglocx,genderlocx,genderlocy
    a=namelocx.value
    b=namelocy.value
    c=agelocx.value
    d=agelocy.value
    e=genderlocx.value
    f=genderlocy.value
    g=symlocx.value
    h=symlocy.value
    i=diaglocx.value
    j=diaglocy.value

    templete = cv2.imread(os.getcwd() + '/templete/prescription-templete.jpg' )
    write(str(addname),(a,b), black)
    write(str(addage),(c,d),black,size=0.45)
    write(str(addgender),(e,f),black)
    write(str(addsignature),(188,765),black)
    write(str(addsymptom),(g,h))
    write(str(adddiagnosis),(i,j))
    write(str(today),(428,274),black,size=0.45)



def write(text,origin, color=black,size=0.6):
    global templete
    cv2.putText(templete, text , origin ,  cv2.FONT_HERSHEY_DUPLEX, size, color , 1, cv2.LINE_AA)

def name(x):
    global addname
    addname = x
    generate()

def age(x):
    global addage
    addage = x
    generate()

def gender(x):
    global addgender
    addgender = x
    generate()

def medicine(value,count,tag):
    global medlocx,doselocx,timelocx
    a=medlocx.value
    b=doselocx.value
    c=timelocx.value
    x = len(value)

    if (tag=="med"):
        for i in range(x):
            cnt = count[i]
            y = value[i]
            z = pos(cnt)
            write(str(y),(a,z),black)
    if (tag == "dose"):
        for i in range(x):
            cnt = count[i]
            y = value[i]
            z = pos(cnt)
            write(str(y), (b, z), black)
    if (tag == "time"):
        for i in range(x):
            cnt = count[i]
            y = value[i]
            z = pos(cnt)
            write(str(y), (c, z), black)


def signature(x):
    global addsignature
    addsignature = sign2.value
    generate()

def symptoms(x):
    global addsymptom
    addsymptom = x
    generate()

def diagnosis(x):
    global adddiagnosis
    adddiagnosis = x
    generate()



def save():
    global addserial
    im_pil = Image.fromarray(cv2.cvtColor(templete,cv2.COLOR_BGR2RGB))
    im_pil.save("prescription" + str(addserial) + ".pdf","PDF",resolution = 100)


