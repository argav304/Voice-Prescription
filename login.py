from tkinter import *
import openpyxl as xl
import mainapp as main
import cv2
from tkinter import messagebox

global namelocx,agelocx,symlocx,diaglocx,genderlocx,medlocx,doselocx,timelocx,namelocy,agelocy,symlocy,diaglocy,genderlocy,medlocy
global count

path = "values.xlsx"
db = xl.load_workbook(path)
table = db.active

login_win=Tk()
login_win.title("Login")
login_win.geometry("400x300")

namelocx = table.cell(row =7 ,column=2)
namelocy= table.cell(row = 7,column=3)
agelocx= table.cell(row = 8,column=2)
agelocy= table.cell(row =8 ,column=3)
genderlocx= table.cell(row =9 ,column=2)
genderlocy= table.cell(row =9 ,column=3)
symlocx= table.cell(row = 10,column=2)
symlocy= table.cell(row = 10,column=3)
diaglocx= table.cell(row = 11,column=2)
diaglocy= table.cell(row = 11,column=3)
medlocx= table.cell(row = 12,column=2)
medlocy= table.cell(row = 12,column=3)
doselocx= table.cell(row = 13,column=2)

timelocx= table.cell(row = 14,column=2)





def register():
    reg_win=Toplevel(login_win)
    reg_win.title("Registration")
    reg_win.geometry("300x500")

    def initialize(x, y):

        global count

        if count == 0:
            namelocx.value, namelocy.value = x, y

        if count == 1:
            agelocx.value, agelocy.value = x, y

        if count == 2:
            genderlocx.value, genderlocy.value = x, y

        if count == 3:
            symlocx.value, symlocy.value = x, y

        if count == 4:
            diaglocx.value, diaglocy.value = x, y

        if count == 5:
            medlocx.value, medlocy.value = x, y

        if count == 6:
            doselocx.value= x

        if count == 7:
            timelocx.value = x

        db.save("values.xlsx")



    def mouse(event,x,y,flags,params):
        global count,namelocx, agelocx, symlocx, diaglocx, genderlocx, medlocx, doselocx, timelocx, namelocy, agelocy, symlocy, diaglocy, genderlocy, medlocy, doselocy, timelocy

        if event == cv2.EVENT_LBUTTONDOWN:
            initialize(x,y)
            count +=1

    def loc_set():
        global count
        count = 0
        messagebox.showinfo("ORDER TO CLICK","SELECT THE ATTRIBUTES IN FOLLOWING ORDER\n Name -> Age -> Gender -> Symptoms-\n->Diagnosis->Medicine->Dose->Time ")
        img = cv2.imread('templete\prescription-templete.jpg')
        cv2.imshow("Original Image",img)
        cv2.setMouseCallback("Original Image",mouse)

        cv2.waitKey(0)




    name1 = table.cell(row=1,column=2)
    med_num1 = table.cell(row=5,column=2)
    username1 = table.cell(row=2,column=2)
    password1 = table.cell(row=3,column=2)
    sign1 = table.cell(row=4,column=2)
    email1 = table.cell(row = 15,column=2)
    epass1 = table.cell(row=16,column=2)


    username = StringVar()
    password = StringVar()
    name = StringVar()
    med_num = StringVar()
    sign = StringVar()
    email = StringVar()
    epass = StringVar()


    Label(reg_win,text="REGISTRATION",font=("Arial", 25)).pack()

    Label(reg_win,text="Name",font=("Arial", 15)).pack(pady=5)
    name_e = Entry(reg_win, textvariable=name)
    name_e.pack()

    Label(reg_win,text="Medical Registration Number",font=("Arial", 15)).pack()
    med_e = Entry(reg_win, textvariable=med_num)
    med_e.pack()

    Label(reg_win,text="Username",font=("Arial", 15)).pack(pady=5)
    user_e = Entry(reg_win, textvariable=username)
    user_e.pack()

    Label(reg_win,text="Password",font=("Arial", 15)).pack()
    pass_e = Entry(reg_win,show="*", textvariable=password)
    pass_e.pack()

    Label(reg_win,text="Signature",font=("Arial", 15)).pack(pady=5)
    sgn_e = Entry(reg_win, textvariable=sign)
    sgn_e.pack()

    Label(reg_win, text="E -Mail", font=("Arial", 15)).pack(pady=5)
    email_e = Entry(reg_win, textvariable=email)
    email_e.pack()

    Label(reg_win, text="Email Password", font=("Arial", 15)).pack(pady=5)
    epass_e = Entry(reg_win,show="*", textvariable=epass)
    epass_e.pack()

    def enter():
        name1.value = name_e.get()
        med_num1.value = med_e.get()
        username1.value = user_e.get()
        password1.value = pass_e.get()
        sign1.value = sgn_e.get()
        email1.value=email_e.get()
        epass1.value=epass_e.get()
        db.save("values.xlsx")



    submit_btn = Button(reg_win,text = "SUBMIT",command = enter).pack(pady=20)



    loc_btn = Button(reg_win,text = "Set Attribute location values", command = loc_set)
    loc_btn.pack(pady = 5)

    loc_btn.config(height = 1,width = 25)



username = table.cell(row = 2, column = 2)
password = table.cell(row = 3, column = 2)



username1 = StringVar()
password1 = StringVar()
#test = StringVar()

bg = PhotoImage(file="bg.png")
label1 = Label( login_win, image=bg)
label1.place(x=0, y=0)



Label(login_win,text="LOGIN",font=("Arial", 25)).pack()

#test_e = Entry(login_win,textvariable=test)
#test_e.pack(side = RIGHT)  #returns "notype" error, if in same line


Label(login_win,text="Username",font=("Arial", 15)).pack(pady=20)
user_e = Entry(login_win, textvariable=username1)
user_e.pack()



Label(login_win,text="Password",font=("Arial", 15)).pack(pady=10)
pass_e = Entry(login_win,show="*", textvariable=password1)
pass_e.pack()



login = PhotoImage(file="login.png")
def validate():
    if ((username.value == user_e.get()) and (password.value == pass_e.get())):
        return True
    else:
        return False
def launch():
    if (validate()):
        login_win.destroy()
        main.mainapp()
    else:
        wrong=Tk()
        wrong.title("wrong details")
        wrong.geometry("100x100")
        Label(wrong,text="wrong details").pack()
        def end():
            wrong.destroy()
        Button(wrong,text="OK",borderwidth=0,command=end).pack()
        wrong.mainloop()
login_btn = Button(login_win,image=login,height = 39, width = 100,borderwidth=0,highlightthickness = 0, bd = 0,command = launch).pack(pady=5)
reg_btn2=Button(login_win,text = "Not registered ? CLick here",borderwidth = 0,command = register).pack()
login_win.mainloop()










