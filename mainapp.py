
from tkinter import *
from tkinter import filedialog
from tkinter import messagebox
import prescription as pr
import speech_recognition as sr
import cv2
import openpyxl as xl
from PIL import Image,ImageTk
import pyttsx3
import os
import smtplib
import base64
from datetime import datetime

from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText





os.environ["TCL_LIBRARY"] = "C:\\Users\\hp\\AppData\\Local\\Programs\\Python\\Python38-32\\tcl\\tcl8.6"
os.environ["TK_LIBRARY"] = "C:\\Users\\hp\\AppData\\Local\\Programs\\Python\\Python38-32\\tcl\\tk8.6"

count=0
dose_count=0
med_count=0
time_count=0
dose_val = []
med_val = []
time_val =[]
dose_countl = []
med_countl = []
time_countl = []

def mainapp():
    path = "patients.xlsx"
    workbook = xl.load_workbook(path)
    table = workbook.active

    path = "values.xlsx"
    db = xl.load_workbook(path)
    table1 = db.active

    email = table1.cell(row=15,column=2)
    epass = table1.cell(row=16,column=2)

    email_val = email.value
    epass_val = epass.value

    global med_val,dose_val,time_val,count,dose_count,med_count,time_count,dose_countl,med_countl,time_countl,datetime
    now = datetime.now()

    date = now.strftime("%d/%m/%Y")
    time = now.strftime("%H:%M:%S")

    window = Tk()
    window.title("VROX")
    window.iconbitmap("ICON vp.ico")
    window.geometry("%dx%d+0+0" % (window.winfo_screenwidth(),window.winfo_screenheight()-100))
    window.resizable(0,0)
    engine=pyttsx3.init()
    voices=engine.getProperty('voices')
    engine.setProperty('voice',voices[1].id)
    r=sr.Recognizer()
    def cv():
        img = cv2.cvtColor(pr.templete, cv2.COLOR_BGR2RGB)
        img = cv2.resize(img,(int(img.shape[1]*(window.winfo_screenheight()/float(img.shape[0]))),window.winfo_screenheight() - 50))
        img = Image.fromarray(img)
        img = ImageTk.PhotoImage(image= img)
        image_frame.ImageTk = img
        image_frame.configure(image = img)
        showid = image_frame.after(10,show)
    def speak(audio):
        engine.say(audio)
        engine.runAndWait()
    engine=pyttsx3.init()
    voices=engine.getProperty('voices')
    engine.setProperty('voice',voices[1].id)
    r=sr.Recognizer()

    left_frame = Frame(window)
    image_frame = Label(left_frame)
    image_frame.grid(row=0,column=0)
    left_frame.grid()
    middle_frame = Frame(window,width =0.700,highlightcolor='green',highlightbackground='green',highlightthickness=1,height=window.winfo_screenheight()-50)
    def name():
            with sr.Microphone() as source:

                    speak("patient\'s name")
                    try:
                            r.adjust_for_ambient_noise(source,duration=0.7)
                            audio= r.listen(source)
                            t1 = r.recognize_google(audio)
                            name_entry.delete(0,END)
                            name_entry.insert(0,t1)
                            pr.name(name_entry.get())
                    except:
                            speak(' could not recognize ')
    name_label = Label(middle_frame,text="Name",font=("Poor Richard",15))
    name_label.grid(row=1,column=0)
    name_entry = Entry(middle_frame,width=50)
    name_entry.grid(row=1,column=1)
    b = Button(middle_frame, text="name",bg='skyblue', command=name)
    b.config( height = 1, width = 6)
    b.grid(row=1,column=5)
    def age():
            with sr.Microphone() as source:

                    speak("patient\'s age")
                    try:
                            r.adjust_for_ambient_noise(source,duration=0.7)
                            audio= r.listen(source)
                            t1 = r.recognize_google(audio)
                            age_entry.delete(0, END)
                            age_entry.insert(0,t1)
                            pr.age(age_entry.get())
                    except:
                            speak(' could not recognize ')
    age_label = Label(middle_frame,text="age",font=("Poor Richard",15))
    age_label.grid(row=5,column=0)
    age_entry = Entry(middle_frame,width=50)
    age_entry.grid(row=5,column=1)
    b1 = Button(middle_frame, text="age",bg='skyblue', command=age)
    b1.config( height = 1, width = 6)
    b1.grid(row=5,column=5)
    def gender():
            with sr.Microphone() as source:
                    speak("patient\'s gender")
                    try:
                            r.adjust_for_ambient_noise(source,duration=0.7)
                            audio= r.listen(source)
                            t1 = r.recognize_google(audio)
                            if t1=='mail':
                                t1='male'
                            gen_entry.delete(0, END)
                            gen_entry.insert(0,t1)
                            pr.gender(gen_entry.get())
                    except:
                            speak(' could not recognize ')
    gen_label = Label(middle_frame,text="gender",font=("Poor Richard",15))
    gen_label.grid(row=10,column=0)
    gen_entry = Entry(middle_frame,width=50)
    gen_entry.grid(row=10,column=1)
    b2 = Button(middle_frame, text="gender",bg='skyblue', command=gender)
    b2.config( height = 1, width = 6)
    b2.grid(row=10,column=5)

#start
    def count_set(flag):
        global med_countl,med_count,dose_countl,dose_count,time_countl,time_count,count
        if (flag == "medicine"):

            med_count+=1
            if(med_count<=count):
                med_count=count
            else:
                count=med_count
            med_countl.append(count)
        if (flag == "dose"):
            dose_count += 1
            if (dose_count <= count):
                dose_count = count
            else:
                count = dose_count
            dose_countl.append(count)
        if (flag == "time"):
            time_count += 1
            if (time_count <= count):
                time_count = count
            else:
                count = time_count
            time_countl.append(count)



    def record_val(widget_val):
        x1 = ""
        with sr.Microphone() as source:
            speak("recite"+widget_val)
            try:
                r.adjust_for_ambient_noise(source, duration=0.7)
                audio = r.listen(source)
                x1 += r.recognize_google(audio)
                if(len(x1)!=0):
                    count_set(widget_val)

            except:
                speak(' could not recognize ')

        return x1

    def delete(flag):
        global count,med_count,dose_count,time_count
        if (flag == "med"):
            if(len(med_val)!=0):
                del med_val[-1]
            if(len(med_countl)!=0):
                del med_countl[-1]
            count-=1
            med_count-=1
        if (flag == "dose"):
            if(len(dose_val)!=0):
                del dose_val[-1]
            if (len(dose_countl) != 0):
                del dose_countl[-1]
            count-=1
            dose_count-=1
        if (flag == "time"):
            if(len(time_val)!=0):            
                del time_val[-1]
            if (len(time_countl) != 0):
                del time_countl[-1]
            count-=1
            time_count-=1
    def entry(tag):
        global count,med_count,time_count,dose_count,med_countl,time_countl,dose_countl
        if tag=="med":
            count_set("medicine")
            
            med_val.append(md_name_entry.get())
        if tag=="dose":
            count_set("dose")
            
            dose_val.append(dose_entry.get())
        if tag=="time":
            count_set("time")
            
            time_val.append(time_entry.get())








    def med_fun(val):       #code to synchronously assign count value to be passed to pr file to print in appropriate row
        global count,dose_count,med_count,time_count,med_countl,med_val,dose_countl,dose_val,time_countl,time_val
        if (val=="medicine"):
            val1 = record_val("medicine")
            med_val.append(val1)




        if (val == "dose"):


            val2 = record_val("dose")
            dose_val.append(val2)



        if (val == "time"):


            val3 = record_val("time")
            time_val.append(val3)

    md_name_entry = Entry(middle_frame,width=35)    #medicine name
    md_name_entry.grid(row=20,column=0)
    b4= Button(middle_frame, text="medicines",bg='skyblue', command = lambda val="medicine" : med_fun(val))
    b4.config( height = 1, width = 8)
    b4.grid(row=25,column=0)
    b4a = Button(middle_frame, text="enter medicine", bg='lightgreen', command = lambda val="med" : entry(val))
    b4a.config(height=1, width=12)
    b4a.grid(row=26,column=0)
    b4b = Button(middle_frame, text="delete medicine", bg='red', command=lambda val="med": delete(val))
    b4b.config(height=1, width=12)
    b4b.grid(row=27, column=0)

    dose_entry = Entry(middle_frame, width=30)      #dose name
    dose_entry.grid(row=20, column=1)
    b41 = Button(middle_frame, text="dose", bg='skyblue', command = lambda val="dose" : med_fun(val))
    b41.config(height=1, width=5)
    b41.grid(row=25, column=1)
    b41a = Button(middle_frame, text="enter dose", bg='lightgreen', command = lambda val="dose" : entry(val))
    b41a.config(height=1, width=8)
    b41a.grid(row=26, column=1)
    b41b = Button(middle_frame, text="delete dose", bg='red', command=lambda val="dose": delete(val))
    b41b.config(height=1, width=8)
    b41b.grid(row=27, column=1)


    time_entry = Entry(middle_frame, width=30)      #time value
    time_entry.grid(row=20, column=2)
    b42 = Button(middle_frame, text="time", bg='skyblue', command = lambda val="time" : med_fun(val))
    b42.config(height=1, width=5)
    b42.grid(row=25, column=2)
    b42a = Button(middle_frame, text="enter time", bg='lightgreen', command = lambda val="time" : entry(val))
    b42a.config(height=1, width=8)
    b42a.grid(row=26, column=2)
    b42b = Button(middle_frame, text="delete time", bg='red', command=lambda val="time": delete(val))
    b42b.config(height=1, width=8)
    b42b.grid(row=27, column=2)

    #end

    def symptoms():
            with sr.Microphone() as source:

                    speak("patient\'s symptoms")
                    try:
                            r.adjust_for_ambient_noise(source,duration=0.7)
                            audio= r.listen(source)
                            t1 = r.recognize_google(audio)
                            sm_entry.delete(0, END)
                            sm_entry.insert(0,t1)
                            pr.symptoms(sm_entry.get())
                    except:
                            speak(' could not recognize ')
    sm_label = Label(middle_frame,text="symptoms",font=("Poor Richard",15))
    sm_label.grid(row=30,column=0)
    sm_entry = Entry(middle_frame,width=50)
    sm_entry.grid(row=30,column=1)
    b5 = Button(middle_frame, text="symptoms",bg='skyblue', command=symptoms)
    b5.config( height = 1, width = 8)
    b5.grid(row=30,column=5)
    def diag():
            with sr.Microphone() as source:

                    speak(" diagnosis ")
                    try:
                            r.adjust_for_ambient_noise(source,duration=0.7)
                            audio= r.listen(source)
                            t1 = r.recognize_google(audio)
                            dia_entry.delete(0, END)
                            dia_entry.insert(0,t1)
                            pr.diagnosis(dia_entry.get())
                    except:
                            speak(' could not recognize ')
    dia_label = Label(middle_frame,text="diagnosis",font=("Poor Richard",15))
    dia_label.grid(row=35,column=0)
    dia_entry = Entry(middle_frame,width=50)
    dia_entry.grid(row=35,column=1)
    b6 = Button(middle_frame, text="diagnosis",bg='skyblue', command=diag)
    b6.config( height = 1, width = 8)
    b6.grid(row=35,column=5)

    sig_label = Label(middle_frame,text="signature",font=("Poor Richard",15))
    sig_label.grid(row=45,column=0)
    sig_entry = Entry(middle_frame,width=50)
    sig_entry.grid(row=45,column=1)

    send2=PhotoImage(file="send.png")
    save2 = PhotoImage(file="save.png")







    def save():

            list = []
            list.append(name_entry.get())
            list.append(date)
            list.append(time)
            table.append(list)
            pr.save()
            pr.serial_set()
            workbook.save("patients.xlsx")


            messagebox.showinfo("information","saved at location : "+os.getcwd())

    b8 = Button(middle_frame,command=save,image=save2,borderwidth=0)
    b8.config( height = 30, width = 60)
    b8.grid(row=50,column=1)

    middle_frame.grid(row=0,column=2,stick=E)

    right_frame = Frame(window,highlightcolor='red',highlightbackground='blue',highlightthickness=1,height=window.winfo_screenheight()-50)

    #e = Label(right_frame,text="your mail id ",font=("Poor Richard",15))
    #e.grid(row=1,column=0)
    #e_entry = Entry(right_frame,width=50)
    #e_entry.grid(row=1,column=1)

    #f= Label(right_frame,text="your password ",font=("Poor Richard",15))
    #f.grid(row=6,column=0)
    #f_entry = Entry(right_frame,width=50)
    #f_entry.config(show='*')
    #f_entry.grid(row=6,column=1)

    g= Label(right_frame,text="host mail id",font=("Poor Richard",15))
    g.grid(row=11,column=0)
    g_entry = Entry(right_frame,width=50)
    g_entry.grid(row=11,column=1)
    right_frame.grid(row=0,column=2,sticky=N)
    def attachments():
        file_path = filedialog.askopenfilename()
        return file_path


    def send():
        try:
            msg = MIMEMultipart()
            filename=attachments()
            with open(filename, "rb") as attachment:
                part = MIMEBase("application", "octet-stream")
                part.set_payload(attachment.read())

            encoders.encode_base64(part)

            part.add_header(
                "Content-Disposition",
                f"attachment; filename= {filename}",)
            msg.attach(part)
            text = msg.as_string()
            server = smtplib.SMTP("smtp.gmail.com", 587)
            server.ehlo()
            server.starttls()
            server.login(email_val,epass_val)
            server.sendmail(email_val,g_entry.get(),text)
            server.close()
            messagebox.showinfo("information","< MAIL SENT >")
        except:
            messagebox.showwarning("warning","mail NOT sent. Try again")

    b9 = Button(right_frame,image=send2,borderwidth=0, command=send)
    b9.config( height = 50, width = 50)
    b9.grid(row=20,column=1)
    def show():

        pr.name(name_entry.get())
        pr.age(age_entry.get())
        pr.gender(gen_entry.get())
        pr.symptoms(sm_entry.get())
        pr.diagnosis(dia_entry.get())
        pr.signature(sig_entry.get())
        pr.medicine(med_val,med_countl,"med")
        pr.medicine(dose_val, dose_countl, "dose")
        pr.medicine(time_val, time_countl, "time")




        cv()
    show()


    def about():
            messagebox.showinfo('VROX','K Srinivasa 18b61a0571\n'
                                                  ' J Dhanush 18b61a0562\n'
                                                  'J Bharadhwa 18b61a0564')
    def guide():
        messagebox.showinfo('USER MANUAL',
            '1.click on the button that u want to fill the field.\n'
                            'eg: name,age,advice,etc...\n'
                            '2.click on save button after completion of writing prescription.\n'
                            '3.enter mail and click on send button to select file and send.')
    def destroy():
        window.destroy()
    menubar = Menu(window)
    window.config(menu=menubar)
    subMenu = Menu(menubar, tearoff=0)
    menubar.add_cascade(label="Help", menu=subMenu)
    subMenu.add_command(label="About", command=about)
    subMenu.add_command(label="guide", command=guide)
    subMenu.add_separator()
    subMenu.add_command(label="Exit", command=destroy)
    window.mainloop()
mainapp()








